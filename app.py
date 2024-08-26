import json
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from utils.constants import Headers


def readFile():
    try:
        inputFilePath = "inputs/data.json"
        file = open(inputFilePath, encoding="utf8")
        data = json.load(file)

        fileData = data["aaData"]

        file.close()

        return fileData

    except:
        raise Exception(
            f"File not found, please check if file is present at {
                inputFilePath}"
        )


def getActions(row):
    html = row["Act"]
    soup = BeautifulSoup(html, "html.parser")

    links = soup.find_all("a")

    listOfActions = list(map(lambda link: link.text.strip(), links))

    return ", ".join(listOfActions)


def getLanguage(row):
    html = row["language"]
    soup = BeautifulSoup(html, "html.parser")

    options = soup.find_all("option")
    language = None
    for option in options:
        if option.has_attr("selected"):
            language = option["value"]
            break

    return language


def getFollowupDateTime(row):
    html = row["Followup_Modify_Date_Time"]
    soup = BeautifulSoup(html, "html.parser")
    return soup.text


def getName(row):
    html = row["name"]
    soup = BeautifulSoup(html, "html.parser")
    return soup.text


def getLeadQuality(row):
    html = row["Lead_Quality"]
    soup = BeautifulSoup(html, "html.parser")
    return soup.find("p").text


def getDescription(row):
    html = row["Description"]
    soup = BeautifulSoup(html, "html.parser")
    times = soup.find_all("b")
    descs = soup.find_all("p")

    descriptionText = ""

    for i, time in enumerate(times):
        descriptionText += f"{time.text} - {descs[i].text}\n"
    return descriptionText


def getPending(row):
    html = row["pending"]
    soup = BeautifulSoup(html, "html.parser")
    return soup.text


def getOriginalDateTime(row):
    val = ''
    try:
        html = row["o_datetime"]
        soup = BeautifulSoup(html, "html.parser")
        val = soup.text

    except:
        print('Error in extracting OriginalDateTime')
        val = 'error'

    return val


def getPainArea(row):
    html = row["painarea"]
    soup = BeautifulSoup(html, "html.parser")
    return soup.find("p").text


def getCourse(row):
    html = row["Course"]
    soup = BeautifulSoup(html, "html.parser")
    options = soup.find_all("option")
    course = "UNALLOCATED"
    for option in options:
        if option.has_attr("selected"):
            course = option["value"]
            break

    return course


# Return json data for the row
def parseRow(row):
    data = {
        # Headers.ACTIONS.value: getActions(row),
        Headers.LANGUAGE.value: getLanguage(row),
        Headers.FOLLOWUP_MODIFY_DATE_TIME.value: getFollowupDateTime(row),
        Headers.NAME.value: getName(row),
        Headers.LEAD_QUALITY.value: getLeadQuality(row),
        Headers.DESCRIPTION.value: getDescription(row),
        Headers.PENDING.value: getPending(row),
        Headers.ORIGINAL_DATETIME.value: getOriginalDateTime(row),
        Headers.PAIN_AREA.value: getPainArea(row),
        Headers.COURSE.value: getCourse(row),
    }

    return data


def configureWB(wb: Workbook, data):
    ws = wb.active

    # Define headers
    headers = [
        # Headers.ACTIONS.value,
        Headers.LANGUAGE.value,
        Headers.FOLLOWUP_MODIFY_DATE_TIME.value,
        Headers.NAME.value,
        Headers.LEAD_QUALITY.value,
        Headers.DESCRIPTION.value,
        Headers.PENDING.value,
        Headers.ORIGINAL_DATETIME.value,
        Headers.PAIN_AREA.value,
        Headers.COURSE.value,
    ]
    ws.append(headers)

    # Define border style
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # default col width
    for column in ws.columns:
        col_letter = column[0].column_letter
        ws.column_dimensions[col_letter].width = 10

    column_widths = {
        "B": 25,
        "C": 15,
        "D": 11,
        "E": 80,
        "F": 16,
        "G": 18,
        "H": 13,
        "I": 18,
    }

    # override custom col width
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Add data rows
    for item in data:
        ws.append([item[header] for header in headers])

    # Apply borders to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border_style

    # Apply text wrapping and bold headers
    for cell in ws["1:1"]:  # Header row
        cell.alignment = Alignment(horizontal="center")
        cell.alignment = Alignment(wrap_text=True)

        cell.font = Font(bold=True)

    # Apply vertical centering to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")  # Center vertically

    # Apply text wrapping to the desired columns
    col_wraps = ["E"]
    for col_letter in col_wraps:
        for cell in ws[col_letter]:
            cell.alignment = Alignment(wrap_text=True)


def generateXls(data):
    try:
        # Create a new workbook
        wb = Workbook()

        configureWB(wb, data)

        # Save the workbook
        outputFolderName = "outputs"
        outputFileName = "output"

        if not os.path.exists(outputFolderName):
            os.makedirs(outputFolderName)

        output_file = outputFolderName + "/" + outputFileName + ".xlsx"

        wb.save(output_file)

        print(f"Excel file '{
              output_file}' created and formatted successfully.")

    except:
        raise Exception("Failed to generate output file")


if __name__ == "__main__":
    rows = readFile()

    jsonData = []

    for row in rows:
        data = parseRow(row)
        jsonData.append(data)
        # break

    generateXls(jsonData)


# if __name__ == "__main__":
#     try:
#         rows = readFile()

#         jsonData = []

#         for row in rows:
#             data = parseRow(row)
#             jsonData.append(data)

#         generateXls(jsonData)

#     except Exception as e:
#         print(e)

#     finally:
#         # To keep terminal window on screen
#         input("Press enter to exit...")
